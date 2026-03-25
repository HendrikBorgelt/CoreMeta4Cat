"""
schema_to_excel.py

Exports the merged CoreMeta4Cat LinkML schema to an Excel workbook with one
sheet per main data class (Synthesis, Characterization, Reaction, Simulation)
plus a summary sheet.

Each sheet uses the same column layout as the original vocabulary reference
workbook so that the output can be compared with the legacy Excel file.

Columns:
  label                      human-readable slot or class name
  parent                     parent slot/class (empty for top-level entries)
  mandatory, recommended, optional  M / R / O classification
  range                      LinkML range type
  slot_uri / class_uri       CURIE for the term
  description                documentation string

Output:
  docs/assets/coremeta4cat_vocabulary.xlsx  (overwrites the file in place)

Run:
    just schema-to-excel
  or directly:
    uv run python scripts/schema_to_excel.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).parent
_ROOT = _HERE.parent
sys.path.insert(0, str(_HERE))

from generate_schema_docs import (  # noqa: E402
    get_all_class_slots,
    get_class_ranged_slot_usage,
    get_slot_details,
    get_subclasses,
    is_mixin,
    load_merged_schema,
    snake_to_readable,
)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "label",
    "parent",
    "mandatory, recommended, optional",
    "range",
    "slot_uri / class_uri",
    "description",
]

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FILL_M = PatternFill("solid", fgColor="FCE4D6")
FILL_R = PatternFill("solid", fgColor="DEEAF1")
FILL_O = PatternFill("solid", fgColor="E2EFDA")


def _slot_mro(schema: dict, class_name: str, slot_name: str) -> str:
    class_def = schema.get("classes", {}).get(class_name, {})
    usage = (class_def.get("slot_usage") or {}).get(slot_name) or {}
    slot_def = get_slot_details(schema, slot_name)
    required    = usage.get("required",    slot_def.get("required",    False))
    recommended = usage.get("recommended", slot_def.get("recommended", False))
    if required:
        return "M"
    if recommended:
        return "R"
    return "O"


def _write_headers(ws) -> None:
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _append_row(ws, row_data: list, mro: str) -> None:
    ws.append(row_data)
    fill = {"M": FILL_M, "R": FILL_R, "O": FILL_O}.get(mro)
    if fill:
        for cell in ws[ws.max_row]:
            cell.fill = fill


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builder
# ─────────────────────────────────────────────────────────────────────────────

def _collect_rows(
    schema: dict,
    class_name: str,
    parent_label: str,
    context_class: str,
    seen: set,
    rows: list,
    depth: int = 0,
) -> None:
    """Recursively collect (label, parent, mro, range, uri, description) tuples."""
    if class_name in seen or depth > 8:
        return
    seen = seen | {class_name}
    classes = schema.get("classes", {})

    direct_slots = get_all_class_slots(schema, class_name)
    for slot_name in direct_slots:
        mro = _slot_mro(schema, context_class if depth == 0 else class_name, slot_name)
        slot_def = get_slot_details(schema, slot_name)
        range_type = slot_def.get("range", "string")
        uri = slot_def.get("slot_uri", "")
        desc = slot_def.get("description", "")
        label = snake_to_readable(slot_name)
        rows.append((label, parent_label, mro, range_type, uri, desc))

        if range_type and range_type in classes and not is_mixin(schema, range_type):
            class_def = classes[range_type]
            class_label = snake_to_readable(range_type)
            class_uri = class_def.get("class_uri", "")
            class_desc = class_def.get("description", "")
            rows.append((class_label, label, mro, "", class_uri, class_desc))
            _collect_rows(schema, range_type, class_label, context_class, seen, rows, depth + 1)
            for sub in get_subclasses(schema, range_type):
                sub_def = classes.get(sub, {})
                sub_label = snake_to_readable(sub)
                sub_uri = sub_def.get("class_uri", "")
                sub_desc = sub_def.get("description", "")
                rows.append((sub_label, class_label, mro, "", sub_uri, sub_desc))
                _collect_rows(schema, sub, sub_label, context_class, seen | {range_type}, rows, depth + 2)

    for su_name, synthetic in get_class_ranged_slot_usage(schema, class_name):
        if su_name in direct_slots:
            continue
        rng = synthetic.get("range", "")
        mro = _slot_mro(schema, class_name, su_name)
        uri = synthetic.get("slot_uri", "")
        desc = synthetic.get("description", "")
        su_label = snake_to_readable(su_name)
        rows.append((su_label, parent_label, mro, rng, uri, desc))
        if rng and rng in classes and not is_mixin(schema, rng):
            class_def = classes[rng]
            class_label = snake_to_readable(rng)
            class_uri = class_def.get("class_uri", "")
            class_desc = class_def.get("description", "")
            rows.append((class_label, su_label, mro, "", class_uri, class_desc))
            _collect_rows(schema, rng, class_label, class_name, seen, rows, depth + 1)
            for sub in get_subclasses(schema, rng):
                sub_def = classes.get(sub, {})
                sub_label = snake_to_readable(sub)
                sub_uri = sub_def.get("class_uri", "")
                sub_desc = sub_def.get("description", "")
                rows.append((sub_label, class_label, mro, "", sub_uri, sub_desc))
                _collect_rows(schema, sub, sub_label, class_name, seen | {rng}, rows, depth + 2)


def build_sheet(wb: openpyxl.Workbook, schema: dict, class_name: str) -> None:
    ws = wb.create_sheet(title=class_name)
    _write_headers(ws)

    rows: list = []
    _collect_rows(schema, class_name, "", class_name, set(), rows)

    # Deduplicate while preserving order
    seen_labels: set = set()
    for row in rows:
        key = (row[0], row[1])
        if key in seen_labels:
            continue
        seen_labels.add(key)
        _append_row(ws, list(row), row[2])

    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main(schema_dir: str, output_path: str) -> None:
    print(f"\nLoading CoreMeta4Cat modules from: {schema_dir}")
    schema = load_merged_schema(schema_dir)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    main_classes = ["Synthesis", "Characterization", "Reaction", "Simulation"]
    for cls in main_classes:
        print(f"  Building sheet: {cls}")
        build_sheet(wb, schema, cls)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    schema_dir = str(_ROOT / "src" / "coremeta4cat" / "schema")
    output_path = str(_ROOT / "docs" / "assets" / "coremeta4cat_vocabulary.xlsx")
    main(schema_dir, output_path)
